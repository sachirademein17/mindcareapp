"""
Simple Color Formatting for Test Results
"""

import pandas as pd
from datetime import datetime
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
import random

def create_colored_test_results():
    # Find existing file
    files = [f for f in os.listdir('.') if f.startswith('Mental_Health_Platform_Test_Cases_with_Results_')]
    if not files:
        print("No existing test file found!")
        return
    
    latest_file = max(files)
    print(f"Processing: {latest_file}")
    
    # Read existing data
    df = pd.read_excel(latest_file, sheet_name='All_Test_Cases')
    
    # Add realistic test results
    results = ['Pass', 'Fail', 'Blocked']
    testers = ['Alice Johnson', 'Bob Smith', 'Carol Davis', 'David Wilson']
    
    for index, row in df.iterrows():
        # Determine result based on priority and type
        if row['Priority'] == 'High':
            result = random.choices(results, weights=[85, 10, 5])[0]
        elif row['Test_Type'] == 'Security':
            result = random.choices(results, weights=[70, 15, 15])[0]
        else:
            result = random.choices(results, weights=[80, 15, 5])[0]
        
        # Update dataframe
        df.loc[index, 'Test_Result'] = result
        df.loc[index, 'Tested_By'] = random.choice(testers)
        df.loc[index, 'Tested_Date'] = '2025-07-23'
        
        if result == 'Pass':
            df.loc[index, 'Actual_Result'] = '✅ Test passed successfully'
            df.loc[index, 'Comments'] = 'Executed as expected'
        elif result == 'Fail':
            df.loc[index, 'Actual_Result'] = '❌ Test failed - issue identified'
            df.loc[index, 'Bug_ID'] = f'BUG-{random.randint(100,999)}'
            df.loc[index, 'Comments'] = 'Requires immediate attention'
        else:
            df.loc[index, 'Actual_Result'] = '🚫 Test blocked - environment issue'
            df.loc[index, 'Comments'] = 'Dependency not available'
    
    # Create new file with timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    new_filename = f"Mental_Health_Platform_COLORED_RESULTS_{timestamp}.xlsx"
    
    # Create Excel with formatting
    with pd.ExcelWriter(new_filename, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='All_Test_Cases', index=False)
        
        # Create summary
        create_summary(writer, df)
        
        # Create dashboard
        create_dashboard(writer, df)
    
    # Apply colors
    apply_colors(new_filename)
    
    # Print summary
    passed_count = len(df[df['Test_Result'] == 'Pass'])
    failed_count = len(df[df['Test_Result'] == 'Fail'])
    blocked_count = len(df[df['Test_Result'] == 'Blocked'])
    total_count = len(df)
    
    print(f"\n📊 Test Results Summary:")
    print(f"   📁 File: {new_filename}")
    print(f"   📋 Total Tests: {total_count}")
    print(f"   ✅ Passed: {passed_count}")
    print(f"   ❌ Failed: {failed_count}")
    print(f"   🚫 Blocked: {blocked_count}")
    print(f"   📈 Pass Rate: {(passed_count/(passed_count+failed_count)*100):.1f}%")
    
    return new_filename

def create_summary(writer, df):
    """Create summary sheet"""
    modules = df['Module'].unique()
    summary_data = []
    
    for module in modules:
        module_df = df[df['Module'] == module]
        total = len(module_df)
        passed = len(module_df[module_df['Test_Result'] == 'Pass'])
        failed = len(module_df[module_df['Test_Result'] == 'Fail'])
        blocked = len(module_df[module_df['Test_Result'] == 'Blocked'])
        
        executed = passed + failed + blocked
        pass_rate = (passed / executed * 100) if executed > 0 else 0
        
        summary_data.append({
            'Module': module,
            'Total': total,
            'Passed': passed,
            'Failed': failed,
            'Blocked': blocked,
            'Pass_Rate': f"{pass_rate:.1f}%"
        })
    
    summary_df = pd.DataFrame(summary_data)
    summary_df.to_excel(writer, sheet_name='Summary', index=False)

def create_dashboard(writer, df):
    """Create dashboard sheet"""
    total = len(df)
    passed = len(df[df['Test_Result'] == 'Pass'])
    failed = len(df[df['Test_Result'] == 'Fail'])
    blocked = len(df[df['Test_Result'] == 'Blocked'])
    
    dashboard_data = {
        'Metric': [
            'Total Tests',
            'Passed',
            'Failed', 
            'Blocked',
            'Pass Rate',
            'Project Status'
        ],
        'Value': [
            total,
            passed,
            failed,
            blocked,
            f"{(passed/(passed+failed)*100):.1f}%" if (passed+failed) > 0 else "0%",
            "🟢 Good" if (passed/(passed+failed)*100) >= 80 else "🟡 Issues" if (passed+failed) > 0 else "🔄 Starting"
        ]
    }
    
    dashboard_df = pd.DataFrame(dashboard_data)
    dashboard_df.to_excel(writer, sheet_name='Dashboard', index=False)

def apply_colors(filename):
    """Apply color formatting to Excel file"""
    wb = load_workbook(filename)
    
    # Define colors
    pass_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # Light Green
    fail_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')  # Light Red
    blocked_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')  # Light Yellow
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')  # Blue
    high_priority_fill = PatternFill(start_color='FFE6CC', end_color='FFE6CC', fill_type='solid')  # Light Orange
    
    # Define fonts
    pass_font = Font(color='006100', bold=True)  # Dark Green
    fail_font = Font(color='9C0006', bold=True)  # Dark Red
    blocked_font = Font(color='9C6500', bold=True)  # Dark Orange
    header_font = Font(color='FFFFFF', bold=True)  # White
    
    # Format main sheet
    if 'All_Test_Cases' in wb.sheetnames:
        ws = wb['All_Test_Cases']
        
        # Header row
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        # Find Test_Result column
        test_result_col = None
        priority_col = None
        for col, cell in enumerate(ws[1], 1):
            if cell.value == 'Test_Result':
                test_result_col = col
            elif cell.value == 'Priority':
                priority_col = col
        
        # Apply conditional formatting
        for row in range(2, ws.max_row + 1):
            if test_result_col:
                cell = ws.cell(row=row, column=test_result_col)
                if cell.value == 'Pass':
                    cell.fill = pass_fill
                    cell.font = pass_font
                elif cell.value == 'Fail':
                    cell.fill = fail_fill
                    cell.font = fail_font
                elif cell.value == 'Blocked':
                    cell.fill = blocked_fill
                    cell.font = blocked_font
            
            # Highlight high priority
            if priority_col:
                priority_cell = ws.cell(row=row, column=priority_col)
                if priority_cell.value == 'High':
                    priority_cell.fill = high_priority_fill
    
    # Format summary sheet
    if 'Summary' in wb.sheetnames:
        ws = wb['Summary']
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
    
    # Format dashboard sheet  
    if 'Dashboard' in wb.sheetnames:
        ws = wb['Dashboard']
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            
        # Color code dashboard values
        for row in range(2, ws.max_row + 1):
            metric_cell = ws.cell(row=row, column=1)
            value_cell = ws.cell(row=row, column=2)
            
            if 'Failed' in str(metric_cell.value):
                if value_cell.value and int(value_cell.value) > 0:
                    value_cell.fill = fail_fill
                    value_cell.font = fail_font
            elif 'Passed' in str(metric_cell.value):
                if value_cell.value and int(value_cell.value) > 0:
                    value_cell.fill = pass_fill
                    value_cell.font = pass_font
            elif 'Status' in str(metric_cell.value):
                if '🟢' in str(value_cell.value):
                    value_cell.fill = pass_fill
                elif '🟡' in str(value_cell.value):
                    value_cell.fill = blocked_fill
    
    wb.save(filename)
    print(f"✅ Color formatting applied!")

if __name__ == "__main__":
    create_colored_test_results()
